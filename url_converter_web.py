import streamlit as st
import pandas as pd
from urllib.parse import urlparse
import re
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import zipfile

# Default project code for naming
DEFAULT_PROJECT_CODE = "GTS2500XX"

# Mapping from locale to region path
LANGUAGE_MAP = {
    "de-DE": "/content/lifetech/europe/en-de",
    "es-ES": "/content/lifetech/europe/en-es",
    "fr-FR": "/content/lifetech/europe/en-fr",
    "ja-JP": "/content/lifetech/japan/en-jp",
    "ko-KR": "/content/lifetech/ipac/en-kr",
    "zh-CN": "/content/lifetech/greater-china/en-cn",
    "zh-TW": "/content/lifetech/ipac/en-tw",
    "pt-BR": "/content/lifetech/latin-america/en-br",
    "es-LATAM": "/content/lifetech/latin-america/en-mx"
}

# --- Functions for Original URL Conversion ---

def clean_url(url):
    if not isinstance(url, str):
        return None
    parsed = urlparse(url)
    path = parsed.path
    if "/home/" not in path:
        return None
    cleaned = path.split("/home/", 1)[1]
    cleaned = re.sub(r'\.html($|[\?#])', r'\1', cleaned)
    return "/home/" + cleaned


def detect_header_row(df):
    for i, row in df.iterrows():
        row_str = row.astype(str).str.lower()
        if any("page title" in cell for cell in row_str) and any("url in aem" in cell for cell in row_str):
            return i
    return 0


def normalize_lang_column(colname):
    match = re.match(r'([a-z]{2}-[A-Z]{2})', str(colname))
    return match.group(1) if match else None


def process_file_original(file):
    df_preview = pd.read_excel(file, sheet_name=0, header=None)
    header_row = detect_header_row(df_preview)
    df = pd.read_excel(file, sheet_name=0, header=header_row)
    df.columns = [str(c).strip().replace("\n", " ") for c in df.columns]

    results = []
    language_columns = {
        col: normalize_lang_column(col)
        for col in df.columns
        if normalize_lang_column(col) in LANGUAGE_MAP
    }

    for _, row in df.iterrows():
        url = next((cell for cell in row if isinstance(cell, str) and "/home/" in cell), None)
        cleaned = clean_url(url)
        if not cleaned:
            continue
        for col, lang in language_columns.items():
            val = row.get(col, "")
            if pd.notna(val) and str(val).strip().lower() in ["x","yes","✓","✔"]:
                results.append({
                    "Original URL": url,
                    "Language": lang,
                    "Localized Path": LANGUAGE_MAP[lang] + cleaned
                })
    return pd.DataFrame(results).sort_values(by=["Language"]) if results else pd.DataFrame()


def style_and_save_excel(df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        ws = writer.sheets['Sheet1']
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 80
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='top')
    buf.seek(0)
    return buf

# --- Product Inclusion List Functions (A### pattern) ---

def detect_first_url_product(row):
    # Try first two columns for a URL
    for idx in [0, 1]:
        try:
            cell = row.iloc[idx]
            if isinstance(cell, str) and cell.lower().startswith('http'):
                return cell
        except:
            pass
    # Fallback: any HTTP cell
    for cell in row:
        if isinstance(cell, str) and cell.lower().startswith('http'):
            return cell
    return None


def process_file_product(file):
    try:
        df = pd.read_excel(file, sheet_name=1, header=2)
    except Exception:
        return pd.DataFrame()
    df.columns = [str(c).strip() for c in df.columns]

    # Identify language columns by header code
    langs = {
        col: re.match(r'([a-z]{2}-[A-Z]{2})', col).group(1)
        for col in df.columns
        if re.match(r'([a-z]{2}-[A-Z]{2})', col) and re.match(r'([a-z]{2}-[A-Z]{2})', col).group(1) in LANGUAGE_MAP
    }

    results = []
    for _, row in df.iterrows():
        pid = None
        # First look for a standalone code cell (A123456)
        for cell in row:
            if isinstance(cell, str) and re.fullmatch(r'A\d{3,6}', cell.strip()):
                pid = cell.strip()
                break
        # If not found, try extracting from the product URL
        if not pid:
            url = detect_first_url_product(row)
            if url:
                parsed = urlparse(url)
                path = parsed.path
                if '/product/' in path:
                    segment = path.split('/product/', 1)[1]
                    pid = segment.split('/')[0].split('.')[0]
                else:
                    m = re.search(r'A\d{3,6}', url)
                    if m:
                        pid = m.group(0)
        if not pid:
            continue

        # Check language markers
        for col, lang in langs.items():
            val = row.get(col, '')
            if pd.notna(val) and str(val).strip().lower() in ['x', 'yes', '✓', '✔']:
                results.append({'Product ID': pid, 'Language': lang})

    return pd.DataFrame(results)


def make_excel_buffer(df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    buf.seek(0)
    return buf

# --- Streamlit UI ---
st.title("🌐 URL Converter Web App")

# GTS ID input for dynamic naming
gts_id = st.text_input("GTS ID", value=DEFAULT_PROJECT_CODE)
project_code = gts_id.strip() if gts_id.strip() else DEFAULT_PROJECT_CODE

uploaded_file = st.file_uploader("Upload an Excel File", type=["xlsx"])
if uploaded_file:
    # Original URL conversion
    df_orig = process_file_original(uploaded_file)
    if df_orig.empty:
        st.warning("No valid data found in the file.")
    else:
        st.success("✅ File processed successfully.")
        st.dataframe(df_orig)
        filename = f"{project_code} - Converted URLs.xlsx"
        st.download_button(
            label="📥 Download Excel",
            data=style_and_save_excel(df_orig),
            file_name=filename
        )

    # Product Inclusion List section
    df_prod = process_file_product(uploaded_file)
    if not df_prod.empty:
        st.header("Product Inclusion List")
        buffers = {}
        for lang in sorted(df_prod['Language'].unique()):
            df_lang = df_prod[df_prod['Language'] == lang][['Product ID']]
            st.subheader(lang)
            st.table(df_lang)
            buf = make_excel_buffer(df_lang)
            st.download_button(
                label="Download Inclusion List",
                data=buf,
                file_name=f"Product Inclusion List_{project_code}_{lang}.xlsx",
                key=f"dl_{lang}"
            )
            buffers[lang] = buf.getvalue()
        # Download all as ZIP
        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, 'w') as zf:
            for lang, data in buffers.items():
                zf.writestr(
                    f"Product Inclusion List_{project_code}_{lang}.xlsx",
                    data
                )
        zip_buf.seek(0)
        st.download_button(
            label="Download All",
            data=zip_buf,
            file_name=f"Product Inclusion Lists_{project_code}.zip",
            mime="application/zip"
        )
